import string
from openpyxl import Workbook


class Spreadsheet:
    """A wrapper around the openpyxl Workbook that provides a typewriter style write-and-advance metaphor"""

    alphabet = list(string.ascii_uppercase)

    def __init__(self):
        self.row = 1
        self.column = 1
        self.workbook = Workbook()
        self.reset()

    def reset(self):
        self.row = 1
        self.column = 1

    def letter(self, index):
        repeat = int(index / len(self.alphabet)) + 1
        index = index % len(self.alphabet)
        return self.alphabet[index] * repeat

    def current_cell(self):
        """
        Return the Excel letter/number coordinates of the current cell using the
        1-based row,column coordinates of the current cell, e.g. 3,2 -> C2
        """
        return '{}{}'.format(self.letter(self.column-1), str(self.row))

    def select_sheet(self, sheet_name):
        """Set the active sheet by the given name (creating it if necessary)"""
        sheet = self.get_sheet(sheet_name)
        self.workbook.active = sheet
        self.reset()

    def get_sheet(self, sheet_name):
        """Return an existing sheet with the given name or create a new sheet with that name"""
        if sheet_name in self.workbook.sheetnames:
            return self.workbook.get_sheet_by_name(sheet_name)
        else:
            return self.workbook.create_sheet(sheet_name)

    def current_sheet(self):
        return self.workbook.active

    def set_value(self, value, advance_row=False, advance_by_rows=1, cell_offset=None):
        self.set_values([value], advance_row=advance_row, advance_by_rows=advance_by_rows, cell_offset=cell_offset)

    def set_values(self, values, advance_row=True, advance_by_rows=1, cell_offset=None):
        """Set the value of the current cell"""
        sheet = self.current_sheet()
        if cell_offset:
            self.column, self.row = cell_offset
        for index, value in enumerate(values):
            cell = self.current_cell()
            sheet[cell] = value
            if index < len(values)-1:
                self.advance_column()
        self.advance(advance_row, advance_by_rows)

    def advance(self, advance_row, advance_by_rows):
        if advance_row or advance_by_rows > 1:
            for _ in range(0, advance_by_rows):
                self.advance_row()
        else:
            self.advance_column()

    def advance_column(self):
        self.column += 1

    def advance_row(self):
        self.column = 1
        self.row += 1

    def save(self, filename):
        self.workbook.save(filename)
